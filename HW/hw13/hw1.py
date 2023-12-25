import openpyxl
def read_excel_files():
    data = []
    for num in range(1, 4):
        file_name = f"{num}{num}{num}{num}.xlsx"
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            wb.close()
        except FileNotFoundError:
            print(f"Файл {file_name} не найден.")
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Файл {file_name} не является файлом Excel.")
    return data

def write_sorted_data_to_excel(data):
    sorted_data = sorted(data, key=lambda row: row[0], reverse=True)
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    for row in sorted_data:
        result_sheet.append(row)
    for row in result_sheet.iter_rows(min_row=2, max_row=result_sheet.max_row):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin"),
                                                 right=openpyxl.styles.Side(border_style="thin"),
                                                 top=openpyxl.styles.Side(border_style="thin"),
                                                 bottom=openpyxl.styles.Side(border_style="thin"))
    result_wb.save("result.xlsx")
    print("Данные записаны в файл result.xlsx")

if __name__ == "__main__":
    data = read_excel_files()
    write_sorted_data_to_excel(data)